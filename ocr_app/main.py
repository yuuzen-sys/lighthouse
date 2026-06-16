"""
FastAPI backend for OCR photo management app.
"""

import os
import re
import json
import shutil
import io
import asyncio
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from datetime import datetime

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from database import init_db, get_conn
from ocr_engine import process_image
from matcher import extract_devices_from_text, extract_prices_from_text
from auth import hash_password, verify_password, create_token, decode_token

_upload_dir = os.getenv("UPLOAD_DIR")
UPLOAD_DIR = Path(_upload_dir) if _upload_dir else Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="Lighthouse")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def _auth_middleware(request: Request, call_next):
    path = request.url.path
    if (path in ("/", "/api/auth/login", "/api/auth/setup", "/api/auth/needs-setup")
            or request.method == "OPTIONS"
            or path.startswith("/static/")
            or path.startswith("/uploads/")):
        return await call_next(request)

    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return JSONResponse(status_code=401, content={"detail": "Not authenticated"})

    username = decode_token(auth_header[7:])
    if not username:
        return JSONResponse(status_code=401, content={"detail": "Token invalid or expired"})

    with get_conn() as conn:
        user = conn.execute(
            "SELECT username FROM users WHERE username = ?", (username,)
        ).fetchone()
    if not user:
        return JSONResponse(status_code=401, content={"detail": "User not found"})

    request.state.username = username
    return await call_next(request)

app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
app.mount("/static", StaticFiles(directory=str(Path(__file__).parent / "frontend")), name="static")

CLAUDE_API_KEY = os.getenv("ANTHROPIC_API_KEY")

# OCR は CPU ヘビーなので 1 スレッドで直列処理（複数写真をキューイング）
_ocr_pool = ThreadPoolExecutor(max_workers=1)


def _do_analysis(photo_id: int):
    """バックグラウンドスレッドで OCR を実行し DB に結果を書く。"""
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM photos WHERE id = ?", (photo_id,)).fetchone()
    if not row:
        return
    photo = dict(row)
    try:
        image_path = str(UPLOAD_DIR / photo["stored_filename"])
        result = process_image(image_path, api_key=CLAUDE_API_KEY)
        device_candidates = extract_devices_from_text(result["raw_text"])
        price_candidates = extract_prices_from_text(result["raw_text"])
        new_status = "analyzed" if device_candidates else "skip"
        with get_conn() as conn:
            conn.execute(
                """INSERT INTO ocr_results
                   (photo_id, engine, confidence, raw_text, device_candidates, price_candidates)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (photo_id, result["engine"], result.get("confidence"),
                 result["raw_text"],
                 json.dumps(device_candidates, ensure_ascii=False),
                 json.dumps(price_candidates, ensure_ascii=False))
            )
            conn.execute(
                "UPDATE photos SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (new_status, photo_id)
            )
    except Exception as exc:
        with get_conn() as conn:
            conn.execute(
                "UPDATE photos SET status='error', note=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (str(exc)[:300], photo_id)
            )


@app.on_event("startup")
def startup():
    init_db()
    _create_initial_admin()


def _create_initial_admin():
    """Create first admin user from ADMIN_USERNAME / ADMIN_PASSWORD env vars if no users exist."""
    uname = os.getenv("ADMIN_USERNAME", "").strip()
    upass = os.getenv("ADMIN_PASSWORD", "").strip()
    if not uname or not upass:
        return
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if count == 0:
            conn.execute(
                "INSERT INTO users (username, hashed_password) VALUES (?, ?)",
                (uname, hash_password(upass))
            )


# ─── Auth ───────────────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    username: str
    password: str


class SetupRequest(BaseModel):
    username: str
    password: str


@app.get("/api/auth/needs-setup")
def auth_needs_setup():
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    return {"needs_setup": count == 0}


@app.post("/api/auth/login")
def auth_login(data: LoginRequest):
    with get_conn() as conn:
        user = conn.execute(
            "SELECT username, hashed_password FROM users WHERE username = ?",
            (data.username,)
        ).fetchone()
    if not user or not verify_password(data.password, user["hashed_password"]):
        raise HTTPException(401, "ユーザー名またはパスワードが正しくありません")
    return {"token": create_token(data.username), "username": data.username}


@app.post("/api/auth/setup")
def auth_setup(data: SetupRequest):
    """Create first user. Only works when no users exist."""
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if count > 0:
            raise HTTPException(400, "セットアップは既に完了しています")
        if not data.username.strip():
            raise HTTPException(400, "ユーザー名が必要です")
        if len(data.password) < 6:
            raise HTTPException(400, "パスワードは6文字以上必要です")
        uname = data.username.strip()
        conn.execute(
            "INSERT INTO users (username, hashed_password) VALUES (?, ?)",
            (uname, hash_password(data.password))
        )
    return {"token": create_token(uname), "username": uname}


@app.get("/api/auth/me")
def auth_me(request: Request):
    return {"username": request.state.username}


# ─── Users ──────────────────────────────────────────────────────────────────

class UserCreate(BaseModel):
    username: str
    password: str


@app.get("/api/users")
def list_users():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT username, created_at FROM users ORDER BY created_at"
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/users")
def create_user(data: UserCreate):
    uname = data.username.strip()
    if not uname:
        raise HTTPException(400, "ユーザー名が必要です")
    if len(data.password) < 6:
        raise HTTPException(400, "パスワードは6文字以上必要です")
    with get_conn() as conn:
        if conn.execute("SELECT 1 FROM users WHERE username = ?", (uname,)).fetchone():
            raise HTTPException(400, "このユーザー名は既に使用されています")
        conn.execute(
            "INSERT INTO users (username, hashed_password) VALUES (?, ?)",
            (uname, hash_password(data.password))
        )
    return {"ok": True}


@app.delete("/api/users/{username}")
def delete_user(username: str, request: Request):
    if username == request.state.username:
        raise HTTPException(400, "自分自身は削除できません")
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if count <= 1:
            raise HTTPException(400, "最後のユーザーは削除できません")
        conn.execute("DELETE FROM users WHERE username = ?", (username,))
    return {"ok": True}


# ─── Root ───────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return FileResponse(Path(__file__).parent / "frontend" / "index.html")


# ─── Stores ─────────────────────────────────────────────────────────────────

_data_dir = os.getenv("DATA_DIR")
_DATA_DIR = Path(_data_dir) if _data_dir else Path(__file__).parent / "data"
CANDIDATES_PATH = _DATA_DIR / "candidates.json"
_BUNDLED_CANDIDATES_PATH = Path(__file__).parent / "data" / "candidates.json"


_DEFAULT_NON_DEVICE_CATS = ["料金表", "店舗外観", "什器", "HR", "その他"]

DEVICE_GROUPS = [
    "最新iPhone", "主要Android", "その他Android",
    "廉価（国産）Android", "廉価（中華）Android", "中古iPhone", "未振り分け",
]

CARRIER_ORDER = ["ドコモ", "au", "SB", "UQ", "Y!", "楽天"]
CARRIER_SYM   = {"ドコモ": "d", "au": "au", "SB": "SB", "UQ": "UQ", "Y!": "Y!", "楽天": "楽天"}

_DEFAULT_RENAME_TAGS: dict[str, str] = {
    "iPhone17_256GB":     "割引POP（I17）",
    "iPhone16e_128GB":    "割引POP（I16e）",
    "iPhone16_128GB":     "割引POP（I16）",
    "iPhone17e_256GB":    "割引POP（I17e）",
    "iPhone17 Pro_256GB": "割引POP（I17Pro）",
    "Google Pixel 9a":    "割引POPA(Google Pixel 9a)",
    "Google Pixel 10":    "割引POPA(Google Pixel 10)",
    "Google Pixel 10a":   "割引POPA(Google Pixel 10a)",
}

def _load_candidates() -> dict:
    if not CANDIDATES_PATH.exists():
        CANDIDATES_PATH.parent.mkdir(parents=True, exist_ok=True)
        bundled = _BUNDLED_CANDIDATES_PATH
        if bundled.exists() and bundled.resolve() != CANDIDATES_PATH.resolve():
            shutil.copy2(str(bundled), str(CANDIDATES_PATH))
        else:
            _save_candidates({
                "devices": {"新品": [], "中古": [], "その他": []},
                "prices": [],
                "non_device_categories": list(_DEFAULT_NON_DEVICE_CATS),
                "device_groups": {},
                "rename_tags": dict(_DEFAULT_RENAME_TAGS),
            })
    with open(CANDIDATES_PATH, encoding='utf-8') as f:
        data = json.load(f)
    dirty = False
    if "non_device_categories" not in data:
        data["non_device_categories"] = list(_DEFAULT_NON_DEVICE_CATS)
        dirty = True
    if "device_groups" not in data:
        data["device_groups"] = {}
        dirty = True
    if "rename_tags" not in data:
        data["rename_tags"] = dict(_DEFAULT_RENAME_TAGS)
        dirty = True
    if dirty:
        _save_candidates(data)
    return data


def _save_candidates(data: dict):
    with open(CANDIDATES_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _parse_store_line(line: str) -> dict | None:
    """Parse 【code】\tname or code,name formats."""
    line = line.strip()
    if not line:
        return None
    # 【DS01】\tドコモショップ... or 【DS01】 ドコモショップ...
    m = re.match(r'【(.+?)】[\t ]+(.+)', line)
    if m:
        return {"code": m.group(1).strip(), "name": m.group(2).strip()}
    # Fallback: code,name
    parts = line.split(',', 1)
    if len(parts) == 2:
        return {"code": parts[0].strip(), "name": parts[1].strip()}
    return {"code": "", "name": line}


class StoreImport(BaseModel):
    text: str


@app.post("/api/stores/import")
def import_stores(data: StoreImport):
    """Import store list. Supports 【code】\\tname or code,name per line."""
    stores = [s for line in data.text.strip().splitlines()
              if (s := _parse_store_line(line))]

    with get_conn() as conn:
        conn.execute("DELETE FROM stores")
        conn.executemany(
            "INSERT INTO stores (code, name) VALUES (:code, :name)",
            stores
        )

    return {"imported": len(stores)}


@app.get("/api/stores")
def list_stores():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM stores ORDER BY code").fetchall()
    return [dict(r) for r in rows]


@app.delete("/api/stores/{store_id}")
def delete_store(store_id: int):
    """Delete a store and all its photos (files + DB records)."""
    with get_conn() as conn:
        store = conn.execute("SELECT * FROM stores WHERE id = ?", (store_id,)).fetchone()
        if not store:
            raise HTTPException(404, "Store not found")

        # Collect all photo IDs (including cropped children)
        photo_rows = conn.execute(
            "SELECT id, stored_filename FROM photos WHERE store_id = ?", (store_id,)
        ).fetchall()
        photo_ids = [r["id"] for r in photo_rows]
        filenames = [r["stored_filename"] for r in photo_rows]

        # Also collect children whose parent belongs to this store
        if photo_ids:
            placeholders = ",".join("?" * len(photo_ids))
            child_rows = conn.execute(
                f"SELECT id, stored_filename FROM photos WHERE parent_id IN ({placeholders})",
                photo_ids
            ).fetchall()
            for r in child_rows:
                if r["id"] not in photo_ids:
                    photo_ids.append(r["id"])
                    filenames.append(r["stored_filename"])

        for pid in photo_ids:
            conn.execute("DELETE FROM ocr_results WHERE photo_id = ?", (pid,))
            conn.execute("DELETE FROM photos WHERE id = ?", (pid,))
        conn.execute("DELETE FROM stores WHERE id = ?", (store_id,))

    for fname in filenames:
        try:
            (UPLOAD_DIR / fname).unlink(missing_ok=True)
        except Exception:
            pass

    return {"deleted_photos": len(filenames), "store_name": store["name"]}


@app.delete("/api/all-data")
def delete_all_data():
    """Delete every store, photo record, OCR result, and uploaded file."""
    with get_conn() as conn:
        rows = conn.execute("SELECT stored_filename FROM photos").fetchall()
        filenames = [r["stored_filename"] for r in rows]
        conn.execute("DELETE FROM ocr_results")
        conn.execute("DELETE FROM photos")
        conn.execute("DELETE FROM stores")

    deleted_files = 0
    for fname in filenames:
        try:
            (UPLOAD_DIR / fname).unlink(missing_ok=True)
            deleted_files += 1
        except Exception:
            pass

    return {"deleted_stores": True, "deleted_photos": len(filenames), "deleted_files": deleted_files}


# ─── Photos ─────────────────────────────────────────────────────────────────

@app.post("/api/photos/upload")
async def upload_photos(
    files: list[UploadFile] = File(...),
    store_id: int = Form(...)
):
    saved = []
    with get_conn() as conn:
        for f in files:
            suffix = Path(f.filename).suffix
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            stored_name = f"{timestamp}{suffix}"
            dest = UPLOAD_DIR / stored_name
            with open(dest, "wb") as out:
                shutil.copyfileobj(f.file, out)

            cursor = conn.execute(
                """INSERT INTO photos (store_id, original_filename, stored_filename)
                   VALUES (?, ?, ?)""",
                (store_id, f.filename, stored_name)
            )
            saved.append({
                "id": cursor.lastrowid,
                "original_filename": f.filename,
                "stored_filename": stored_name,
            })
    return {"uploaded": len(saved), "photos": saved}


@app.get("/api/photos")
def list_photos(store_id: int | None = None):
    with get_conn() as conn:
        if store_id:
            rows = conn.execute(
                """SELECT p.*, s.code as store_code, s.name as store_name
                   FROM photos p LEFT JOIN stores s ON p.store_id = s.id
                   WHERE p.store_id = ? ORDER BY p.created_at""",
                (store_id,)
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT p.*, s.code as store_code, s.name as store_name
                   FROM photos p LEFT JOIN stores s ON p.store_id = s.id
                   ORDER BY p.store_id, p.created_at"""
            ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        try:
            d["carriers"] = json.loads(d["carriers"]) if d.get("carriers") else []
        except Exception:
            d["carriers"] = []
        try:
            d["contract_types"] = json.loads(d["contract_types"]) if d.get("contract_types") else []
        except Exception:
            d["contract_types"] = []
        result.append(d)
    return result


@app.post("/api/photos/{photo_id}/analyze")
async def analyze_photo(photo_id: int):
    """OCR をバックグラウンドで開始し即座に返る。進捗は status='analyzing' で追跡。"""
    with get_conn() as conn:
        photo = conn.execute("SELECT * FROM photos WHERE id = ?", (photo_id,)).fetchone()
        if not photo:
            raise HTTPException(404, "Photo not found")
        conn.execute(
            "UPDATE photos SET status='analyzing', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (photo_id,)
        )
    loop = asyncio.get_event_loop()
    loop.run_in_executor(_ocr_pool, _do_analysis, photo_id)
    return {"status": "analyzing"}


@app.get("/api/photos/{photo_id}/ocr-result")
def get_ocr_result(photo_id: int):
    """最新の OCR 結果を返す（解析完了後に取得するため）。"""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ocr_results WHERE photo_id=? ORDER BY created_at DESC LIMIT 1",
            (photo_id,)
        ).fetchone()
    if not row:
        return {"found": False}
    r = dict(row)
    return {
        "found": True,
        "engine": r["engine"],
        "confidence": r["confidence"],
        "raw_text": r["raw_text"],
        "device_candidates": json.loads(r["device_candidates"] or "[]"),
        "price_candidates": json.loads(r["price_candidates"] or "[]"),
    }


class PhotoUpdate(BaseModel):
    device_name: str | None = None
    device_category: str | None = None
    price: str | None = None
    note: str | None = None
    carriers: list[str] | None = None
    contract_types: list[str] | None = None
    non_device_category: str | None = None


def _auto_register_device(name: str, category: str | None):
    """Add device to candidates list if not already present."""
    if not name:
        return
    data = _load_candidates()
    cat = category or "その他"
    devices = data["devices"].setdefault(cat, [])
    if name not in devices:
        devices.append(name)
        _save_candidates(data)


_WIN_INVALID = str.maketrans({
    "\\": "_", "/": "_", ":": "：", "*": "＊",
    "?": "？", '"': "'", "<": "＜", ">": "＞", "|": "｜",
})


def _sanitize_filename(name: str) -> str:
    return name.translate(_WIN_INVALID).strip()


def _build_renamed(code: str, store_name: str, note: str, device_name: str, suffix: str) -> str:
    """Legacy fallback: 【code】店舗名 備考(機種名).ext"""
    prefix = f"【{code}】{store_name}" if code else (store_name or "")
    note   = (note or "").strip()
    device = (device_name or "").strip()
    if note and device:
        body = f"{note}({device})"
    elif device:
        body = f"({device})"
    else:
        body = note
    full = f"{prefix} {body}".strip() if body else prefix
    return _sanitize_filename(full) + suffix


def _pop_tag(device: str, price: str, non_dev_cat: str,
             rename_tags: dict, device_groups: dict) -> str:
    """Return the POP tag (without carrier prefix) for a photo."""
    if non_dev_cat:
        return f"POP({non_dev_cat})"
    if not device:
        return ""
    if price and device in rename_tags:
        return rename_tags[device]
    grp = device_groups.get(device, "未振り分け")
    if "iPhone" in grp:
        return "中古端末割引POPI" if grp == "中古iPhone" else "POPI"
    if grp == "廉価（中華）Android":
        return "割引POP中華A"
    return "POPA"


def _build_renamed_new(code: str, store_name: str,
                       device: str, non_dev_cat: str, price: str,
                       carriers: list[str], contracts: list[str],
                       rename_tags: dict, device_groups: dict,
                       suffix: str) -> str:
    """
    【code】店舗名 [d/au/SB/UQ/Y!/楽天][タグ] ... [新規契約あり].ext
    キャリア順: ドコモ > au > SB > UQ > Y! > 楽天
    """
    prefix = f"【{code}】{store_name}" if code else (store_name or "")
    tag = _pop_tag(device, price, non_dev_cat, rename_tags, device_groups)
    if not tag:
        return _sanitize_filename(prefix) + suffix
    if carriers:
        ordered = [c for c in CARRIER_ORDER if c in set(carriers)]
        parts   = [f"{CARRIER_SYM[c]}{tag}" for c in ordered]
    else:
        parts = [tag]
    body = " ".join(parts)
    if "新規" in (contracts or []):
        body += " 新規契約あり"
    full = f"{prefix} {body}".strip()
    return _sanitize_filename(full) + suffix


@app.put("/api/photos/{photo_id}")
def update_photo(photo_id: int, data: PhotoUpdate):
    """Save confirmed device name/price. Auto-registers new device names to candidates."""
    if data.device_name:
        _auto_register_device(data.device_name, data.device_category)

    carriers_json       = json.dumps(data.carriers,       ensure_ascii=False) if data.carriers       is not None else None
    contract_types_json = json.dumps(data.contract_types, ensure_ascii=False) if data.contract_types is not None else None
    with get_conn() as conn:
        conn.execute(
            """UPDATE photos
               SET device_name = ?, device_category = ?, price = ?, note = ?,
                   carriers = ?, contract_types = ?, non_device_category = ?,
                   status = 'confirmed', updated_at = CURRENT_TIMESTAMP
               WHERE id = ?""",
            (data.device_name, data.device_category, data.price, data.note,
             carriers_json, contract_types_json, data.non_device_category, photo_id)
        )
    return {"ok": True}


class ConfirmRequest(BaseModel):
    renamed_filename: str | None = None


@app.post("/api/photos/{photo_id}/confirm")
def confirm_photo(photo_id: int, req: ConfirmRequest = ConfirmRequest()):
    """Finalize photo: generate renamed filename using carrier/tag rules."""
    with get_conn() as conn:
        row = conn.execute(
            """SELECT p.*, s.code as store_code, s.name as store_name
               FROM photos p LEFT JOIN stores s ON p.store_id = s.id
               WHERE p.id = ?""",
            (photo_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "Photo not found")

    photo     = dict(row)
    code      = photo.get("store_code") or ""
    store     = photo.get("store_name") or ""
    device    = photo.get("device_name") or ""
    non_dev   = photo.get("non_device_category") or ""
    price     = photo.get("price") or ""
    carriers  = json.loads(photo.get("carriers")  or "[]")
    contracts = json.loads(photo.get("contract_types") or "[]")
    suffix    = Path(photo["original_filename"]).suffix or ".jpg"

    cands        = _load_candidates()
    rename_tags  = cands.get("rename_tags", {})
    device_groups = cands.get("device_groups", {})

    if req.renamed_filename:
        renamed = _sanitize_filename(req.renamed_filename) + suffix
    else:
        renamed = _build_renamed_new(
            code, store, device, non_dev, price,
            carriers, contracts, rename_tags, device_groups, suffix,
        )

    with get_conn() as conn:
        conn.execute(
            "UPDATE photos SET renamed_filename = ?, status = 'done', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (renamed, photo_id)
        )

    return {"renamed_filename": renamed}


class PerspPoint(BaseModel):
    x: int
    y: int


class PerspCropRequest(BaseModel):
    points: list[PerspPoint]
    force_ratio: bool = False   # True = 透視補正後に 1.41:1 へリサイズ


@app.post("/api/photos/{photo_id}/perspective-crop")
async def perspective_crop_photo(photo_id: int, req: PerspCropRequest):
    """
    4点透視補正でトリミング。歪んだPOPを長方形に補正して保存。
    points順: 左上→右上→右下→左下
    force_ratio=True のとき、長辺:短辺 = 1.41:1 にリサイズして保存。
    """
    import cv2
    import numpy as np

    if len(req.points) != 4:
        raise HTTPException(400, "4点必要です")

    with get_conn() as conn:
        photo = conn.execute("SELECT * FROM photos WHERE id = ?", (photo_id,)).fetchone()
        if not photo:
            raise HTTPException(404, "Photo not found")

    source_path = UPLOAD_DIR / photo["stored_filename"]
    with open(source_path, "rb") as f:
        arr = np.frombuffer(f.read(), dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        raise HTTPException(500, "Cannot decode image")

    src = np.float32([[p.x, p.y] for p in req.points])
    w1 = float(np.linalg.norm(src[1] - src[0]))
    w2 = float(np.linalg.norm(src[2] - src[3]))
    h1 = float(np.linalg.norm(src[3] - src[0]))
    h2 = float(np.linalg.norm(src[2] - src[1]))
    out_w = max(1, int(max(w1, w2)))
    out_h = max(1, int(max(h1, h2)))

    dst = np.float32([[0, 0], [out_w - 1, 0], [out_w - 1, out_h - 1], [0, out_h - 1]])
    M = cv2.getPerspectiveTransform(src, dst)
    corrected = cv2.warpPerspective(img, M, (out_w, out_h))

    # 縦横比 1.41:1 補正
    # ①→②(上辺) を常に短辺（幅）、②→③(右辺) を常に長辺（高さ）として固定する。
    # 角度が強い場合に距離の大小が逆転しても向きが狂わないようにするため、
    # 距離比較による判定は行わない。
    if req.force_ratio:
        new_w = out_w
        new_h = max(1, round(out_w * 1.41))
        corrected = cv2.resize(corrected, (new_w, new_h), interpolation=cv2.INTER_LANCZOS4)

    suffix = Path(photo["original_filename"]).suffix.lower() or ".jpg"
    base_name = Path(photo["original_filename"]).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    stored_name = f"{timestamp}_persp{suffix}"
    orig_name = f"{base_name}_persp{suffix}"

    ok, buf = cv2.imencode(suffix, corrected)
    if not ok:
        raise HTTPException(500, "Encode failed")
    with open(UPLOAD_DIR / stored_name, "wb") as f:
        f.write(buf.tobytes())

    with get_conn() as conn:
        cursor = conn.execute(
            """INSERT INTO photos (store_id, parent_id, original_filename, stored_filename)
               VALUES (?, ?, ?, ?)""",
            (photo["store_id"], photo_id, orig_name, stored_name)
        )
    return {"id": cursor.lastrowid, "filename": orig_name}


class RotateRequest(BaseModel):
    degrees: int  # 90=右(時計回り), -90=左(反時計回り), 180


@app.post("/api/photos/{photo_id}/rotate")
def rotate_photo(photo_id: int, req: RotateRequest):
    """
    Rotate and overwrite the stored image file.
    Applies EXIF normalization first so the result matches what the user sees.
    """
    from PIL import Image as PILImage, ImageOps

    if req.degrees not in (90, -90, 180, 270, -270):
        raise HTTPException(400, "degrees は 90 / -90 / 180 のいずれかを指定してください")

    with get_conn() as conn:
        photo = conn.execute("SELECT * FROM photos WHERE id = ?", (photo_id,)).fetchone()
        if not photo:
            raise HTTPException(404, "Photo not found")

    path   = UPLOAD_DIR / photo["stored_filename"]
    suffix = path.suffix.lower()
    fmt    = "JPEG" if suffix in (".jpg", ".jpeg") else "PNG"

    with open(path, "rb") as f:
        img = PILImage.open(io.BytesIO(f.read()))
        img.load()

    # EXIF の向き情報を画素に焼き込んでから回転する
    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass

    # Pillow は正数が反時計回りなので符号を反転
    rotated = img.rotate(-req.degrees, expand=True)

    buf = io.BytesIO()
    if fmt == "JPEG":
        rotated.save(buf, format=fmt, quality=95)
    else:
        rotated.save(buf, format=fmt)

    with open(path, "wb") as f:
        f.write(buf.getvalue())

    return {"ok": True}


@app.post("/api/photos/{photo_id}/skip")
def skip_photo(photo_id: int):
    with get_conn() as conn:
        conn.execute(
            "UPDATE photos SET status='skip', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (photo_id,)
        )
    return {"ok": True}


@app.post("/api/photos/{photo_id}/unskip")
def unskip_photo(photo_id: int):
    with get_conn() as conn:
        conn.execute(
            "UPDATE photos SET status='pending', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (photo_id,)
        )
    return {"ok": True}


@app.delete("/api/photos/{photo_id}")
def delete_photo(photo_id: int):
    """
    Delete a photo and its direct children (crops).
    Removes DB records, OCR results, and the actual files from disk.
    """
    with get_conn() as conn:
        # Collect the target photo + all its children
        targets = [photo_id]
        children = conn.execute(
            "SELECT id FROM photos WHERE parent_id = ?", (photo_id,)
        ).fetchall()
        targets.extend(r["id"] for r in children)

        filenames = []
        for pid in targets:
            row = conn.execute(
                "SELECT stored_filename FROM photos WHERE id = ?", (pid,)
            ).fetchone()
            if row:
                filenames.append(row["stored_filename"])
            conn.execute("DELETE FROM ocr_results WHERE photo_id = ?", (pid,))
            conn.execute("DELETE FROM photos WHERE id = ?", (pid,))

    # Delete files from disk (after DB commit)
    for fname in filenames:
        path = UPLOAD_DIR / fname
        try:
            path.unlink(missing_ok=True)
        except Exception:
            pass

    return {"deleted": len(targets)}


# ─── Crop ────────────────────────────────────────────────────────────────────

class CropRegion(BaseModel):
    x: int
    y: int
    w: int
    h: int


@app.post("/api/photos/{photo_id}/crop")
def crop_photo(photo_id: int, regions: list[CropRegion]):
    """
    Crop a photo into multiple regions.
    Each region is saved as a new pending photo (child of the original).
    """
    from PIL import Image as PILImage

    with get_conn() as conn:
        photo = conn.execute("SELECT * FROM photos WHERE id = ?", (photo_id,)).fetchone()
        if not photo:
            raise HTTPException(404, "Photo not found")

    source_path = UPLOAD_DIR / photo["stored_filename"]
    suffix = Path(photo["original_filename"]).suffix.lower() or ".jpg"
    base_name = Path(photo["original_filename"]).stem
    fmt = "JPEG" if suffix in (".jpg", ".jpeg") else "PNG"

    # Load image via bytes to avoid Japanese path issues with PIL
    with open(source_path, "rb") as f:
        img = PILImage.open(io.BytesIO(f.read()))
        img.load()

    created = []
    with get_conn() as conn:
        for i, region in enumerate(regions, 1):
            x1 = max(0, region.x)
            y1 = max(0, region.y)
            x2 = min(img.width, region.x + region.w)
            y2 = min(img.height, region.y + region.h)
            if x2 <= x1 or y2 <= y1:
                continue

            cropped = img.crop((x1, y1, x2, y2))
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            stored_name = f"{timestamp}_crop{i}{suffix}"

            buf = io.BytesIO()
            cropped.save(buf, format=fmt)
            with open(UPLOAD_DIR / stored_name, "wb") as f:
                f.write(buf.getvalue())

            orig_name = f"{base_name}_crop{i}{suffix}"
            cursor = conn.execute(
                """INSERT INTO photos (store_id, parent_id, original_filename, stored_filename)
                   VALUES (?, ?, ?, ?)""",
                (photo["store_id"], photo_id, orig_name, stored_name)
            )
            created.append({"id": cursor.lastrowid, "filename": orig_name})

    return {"created": len(created), "photos": created}


# ─── Candidates ─────────────────────────────────────────────────────────────

@app.get("/api/candidates")
def get_candidates():
    return _load_candidates()


@app.put("/api/candidates")
def update_candidates(data: dict):
    _save_candidates(data)
    return {"ok": True}


class DeviceAdd(BaseModel):
    name: str
    category: str


@app.post("/api/candidates/device")
def add_device_candidate(data: DeviceAdd):
    """Add a single device to a category."""
    cands = _load_candidates()
    devices = cands["devices"].setdefault(data.category, [])
    if data.name not in devices:
        devices.append(data.name)
        _save_candidates(cands)
        return {"added": True}
    return {"added": False, "reason": "already exists"}


class DeviceDelete(BaseModel):
    name: str
    category: str


@app.delete("/api/candidates/device")
def delete_device_candidate(data: DeviceDelete):
    """Remove a device from a category."""
    cands = _load_candidates()
    devices = cands["devices"].get(data.category, [])
    if data.name in devices:
        devices.remove(data.name)
        _save_candidates(cands)
        return {"removed": True}
    return {"removed": False}


class CategoryNameBody(BaseModel):
    name: str


@app.post("/api/candidates/non-device-category")
def add_non_device_category(data: CategoryNameBody):
    cands = _load_candidates()
    cats = cands.setdefault("non_device_categories", list(_DEFAULT_NON_DEVICE_CATS))
    if data.name not in cats:
        cats.append(data.name)
        _save_candidates(cands)
        return {"added": True}
    return {"added": False, "reason": "already exists"}


@app.delete("/api/candidates/non-device-category")
def delete_non_device_category(data: CategoryNameBody):
    cands = _load_candidates()
    cats = cands.get("non_device_categories", [])
    if data.name in cats:
        cats.remove(data.name)
        _save_candidates(cands)
        return {"removed": True}
    return {"removed": False}


class RenameTagBody(BaseModel):
    device: str
    tag: str


class RenameTagDeleteBody(BaseModel):
    device: str


@app.post("/api/candidates/rename-tag")
def set_rename_tag(data: RenameTagBody):
    """Add or update a 1:1 device→rename-tag mapping."""
    cands = _load_candidates()
    cands.setdefault("rename_tags", {})[data.device] = data.tag
    _save_candidates(cands)
    return {"ok": True}


@app.delete("/api/candidates/rename-tag")
def delete_rename_tag(data: RenameTagDeleteBody):
    """Remove a device rename-tag mapping."""
    cands = _load_candidates()
    tags  = cands.get("rename_tags", {})
    if data.device in tags:
        del tags[data.device]
        _save_candidates(cands)
        return {"removed": True}
    return {"removed": False}


class DeviceGroupBody(BaseModel):
    device: str
    group: str


@app.post("/api/candidates/device-group")
def set_device_group(data: DeviceGroupBody):
    """Assign a device to a device group (drag & drop in settings)."""
    if data.group not in DEVICE_GROUPS:
        return {"ok": False, "reason": "unknown group"}
    cands = _load_candidates()
    groups = cands.setdefault("device_groups", {})
    groups[data.device] = data.group
    _save_candidates(cands)
    return {"ok": True}


# ─── Export ─────────────────────────────────────────────────────────────────

@app.get("/api/export")
def export_excel(store_id: int | None = None):
    """Export photo list as Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    with get_conn() as conn:
        if store_id:
            rows = conn.execute(
                """SELECT p.*, s.code as store_code, s.name as store_name
                   FROM photos p LEFT JOIN stores s ON p.store_id = s.id
                   WHERE p.store_id = ? ORDER BY p.created_at""",
                (store_id,)
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT p.*, s.code as store_code, s.name as store_name
                   FROM photos p LEFT JOIN stores s ON p.store_id = s.id
                   ORDER BY p.store_id, p.created_at"""
            ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "写真一覧"

    headers = ["店舗コード", "店舗名", "元ファイル名", "機種名", "カテゴリ", "端末以外区分", "キャリア", "契約種別", "金額", "備考", "リネーム後ファイル名", "ステータス"]
    header_fill = PatternFill(start_color="2B5799", end_color="2B5799", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(rows, 2):
        r = dict(r)
        try:
            carriers_val = "、".join(json.loads(r.get("carriers") or "[]"))
        except Exception:
            carriers_val = ""
        try:
            contracts_val = "、".join(json.loads(r.get("contract_types") or "[]"))
        except Exception:
            contracts_val = ""
        ws.cell(row=row_idx, column=1,  value=r.get("store_code", ""))
        ws.cell(row=row_idx, column=2,  value=r.get("store_name", ""))
        ws.cell(row=row_idx, column=3,  value=r.get("original_filename", ""))
        ws.cell(row=row_idx, column=4,  value=r.get("device_name", ""))
        ws.cell(row=row_idx, column=5,  value=r.get("device_category", ""))
        ws.cell(row=row_idx, column=6,  value=r.get("non_device_category", ""))
        ws.cell(row=row_idx, column=7,  value=carriers_val)
        ws.cell(row=row_idx, column=8,  value=contracts_val)
        ws.cell(row=row_idx, column=9,  value=r.get("price", ""))
        ws.cell(row=row_idx, column=10, value=r.get("note", ""))
        ws.cell(row=row_idx, column=11, value=r.get("renamed_filename", ""))
        ws.cell(row=row_idx, column=12, value=r.get("status", ""))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    filename = f"photos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Devices ─────────────────────────────────────────────────────────────────

@app.get("/api/devices")
def list_devices():
    """Distinct device names that appear in any photo."""
    with get_conn() as conn:
        rows = conn.execute(
            """SELECT DISTINCT device_name FROM photos
               WHERE device_name IS NOT NULL AND device_name != ''
               ORDER BY device_name"""
        ).fetchall()
    return [r["device_name"] for r in rows]


# ─── File Export (server-side copy) ──────────────────────────────────────────

def _safe_copy(src: Path, dest_dir: Path, fname: str) -> str:
    """Copy src → dest_dir/fname; append (2),(3)... on name collision."""
    stem   = Path(fname).stem
    ext    = Path(fname).suffix
    target = dest_dir / fname
    n = 1
    while target.exists():
        n += 1
        target = dest_dir / f"{stem} ({n}){ext}"
    shutil.copy2(str(src), str(target))
    return target.name


class FileExportRequest(BaseModel):
    dest_folder: str
    store_id: int | None = None
    device_name: str | None = None


@app.get("/api/export/zip")
def export_zip(mode: str = "confirmed", store_id: int | None = None, device_name: str | None = None):
    """Stream a ZIP of photos to the browser. mode=confirmed|unprocessed."""
    import zipfile

    if mode == "unprocessed":
        q = "SELECT stored_filename, original_filename FROM photos WHERE status IN ('pending','skip','error')"
        params: list = []
        if store_id:
            q += " AND store_id = ?"; params.append(store_id)
    else:
        q = """SELECT p.stored_filename, p.renamed_filename, p.original_filename,
                      s.code as store_code, s.name as store_name,
                      p.device_name, p.note, p.non_device_category, p.price
               FROM photos p LEFT JOIN stores s ON p.store_id = s.id
               WHERE p.status = 'done'"""
        params = []
        if store_id:
            q += " AND p.store_id = ?"; params.append(store_id)
        if device_name:
            q += " AND p.device_name = ?"; params.append(device_name)

    with get_conn() as conn:
        rows = conn.execute(q, params).fetchall()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        seen: dict[str, int] = {}
        for row in rows:
            r = dict(row)
            src = UPLOAD_DIR / r["stored_filename"]
            if not src.exists():
                continue

            orig_suffix = Path(r["stored_filename"]).suffix.lower() or ".jpg"
            if mode == "unprocessed":
                fname = _sanitize_filename(Path(r["original_filename"]).stem) + Path(r["original_filename"]).suffix
            elif r.get("renamed_filename"):
                base = Path(r["renamed_filename"])
                fname = _sanitize_filename(base.stem) + (base.suffix or orig_suffix)
            else:
                fname = _build_renamed(
                    r.get("store_code") or "", r.get("store_name") or "",
                    r.get("note") or "", r.get("device_name") or "", orig_suffix,
                )

            # Deduplicate names inside the ZIP
            if fname in seen:
                seen[fname] += 1
                stem, ext = Path(fname).stem, Path(fname).suffix
                fname = f"{stem} ({seen[fname]}){ext}"
            else:
                seen[fname] = 1

            with open(src, "rb") as f:
                zf.writestr(fname, f.read())

    zip_name = f"photos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_name}"'},
    )


@app.get("/api/export/count")
def export_count(mode: str, store_id: int | None = None, device_name: str | None = None):
    """Return how many photos would be exported for the given parameters."""
    with get_conn() as conn:
        if mode == "unprocessed":
            q = "SELECT COUNT(*) as c FROM photos WHERE status IN ('pending','skip','error')"
            p: list = []
            if store_id:
                q += " AND store_id = ?"; p.append(store_id)
        else:
            q = "SELECT COUNT(*) as c FROM photos WHERE status = 'done'"
            p = []
            if store_id:
                q += " AND store_id = ?";   p.append(store_id)
            if device_name:
                q += " AND device_name = ?"; p.append(device_name)
        row = conn.execute(q, p).fetchone()
    return {"count": row["c"]}


@app.post("/api/export/confirmed")
def export_confirmed(req: FileExportRequest):
    """Copy status='done' photos to dest_folder using their renamed filenames."""
    dest = Path(req.dest_folder)
    try:
        dest.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise HTTPException(400, f"フォルダ作成失敗: {e}")

    qparts = ["""
        SELECT p.*, s.code as store_code, s.name as store_name
        FROM photos p LEFT JOIN stores s ON p.store_id = s.id
        WHERE p.status = 'done'
    """]
    params: list = []
    if req.store_id:
        qparts.append("AND p.store_id = ?");   params.append(req.store_id)
    if req.device_name:
        qparts.append("AND p.device_name = ?"); params.append(req.device_name)
    qparts.append("ORDER BY p.store_id, p.created_at")

    with get_conn() as conn:
        rows = conn.execute(" ".join(qparts), params).fetchall()

    copied = skipped = 0
    for row in rows:
        ph  = dict(row)
        src = UPLOAD_DIR / ph["stored_filename"]
        if not src.exists():
            skipped += 1; continue

        ext = Path(ph["stored_filename"]).suffix.lower() or ".jpg"
        if ph.get("renamed_filename"):
            base = Path(ph["renamed_filename"])
            fname = _sanitize_filename(base.stem) + (base.suffix or ext)
        else:
            fname = _build_renamed(
                ph.get("store_code") or "",
                ph.get("store_name") or "",
                ph.get("note") or "",
                ph.get("device_name") or "",
                ext,
            )
        _safe_copy(src, dest, fname)
        copied += 1

    return {"copied": copied, "skipped": skipped, "dest": str(dest)}


@app.post("/api/export/unprocessed")
def export_unprocessed(req: FileExportRequest):
    """Copy pending/skip/error photos (original filenames) to dest_folder."""
    dest = Path(req.dest_folder)
    try:
        dest.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise HTTPException(400, f"フォルダ作成失敗: {e}")

    qparts = ["SELECT * FROM photos WHERE status IN ('pending','skip','error')"]
    params: list = []
    if req.store_id:
        qparts.append("AND store_id = ?"); params.append(req.store_id)
    qparts.append("ORDER BY created_at")

    with get_conn() as conn:
        rows = conn.execute(" ".join(qparts), params).fetchall()

    copied = skipped = 0
    for row in rows:
        ph  = dict(row)
        src = UPLOAD_DIR / ph["stored_filename"]
        if not src.exists():
            skipped += 1; continue
        orig = ph["original_filename"]
        fname = _sanitize_filename(Path(orig).stem) + Path(orig).suffix
        _safe_copy(src, dest, fname)
        copied += 1

    return {"copied": copied, "skipped": skipped, "dest": str(dest)}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
